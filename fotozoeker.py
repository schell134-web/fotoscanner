import os
import tkinter as tk
from tkinter import filedialog, messagebox, Scrollbar, Canvas, Frame
from PIL import Image, ImageTk
from openpyxl import load_workbook


# ====== CONFIG ======
# Kolommen (1-based index): B = 2, C = 3
COLUMN_FILE = 2
COLUMN_PERSON = 3
# ====================


class PhotoSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Fotozoeker op persoon (Excel)")

        # Excel-pad + persoon invoer
        top_frame = tk.Frame(root)
        top_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(top_frame, text="Excel-bestand:").grid(row=0, column=0, sticky="w")
        self.excel_entry = tk.Entry(top_frame, width=50)
        self.excel_entry.grid(row=0, column=1, padx=5)
        tk.Button(top_frame, text="Bladeren...", command=self.browse_excel).grid(row=0, column=2, padx=5)

        tk.Label(top_frame, text="Persoonsnaam:").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.person_entry = tk.Entry(top_frame, width=30)
        self.person_entry.grid(row=1, column=1, padx=5, pady=(5, 0), sticky="w")


        tk.Label(top_frame, text="Categorie:").grid(row=2, column=0, sticky="w", pady=(5, 0))
        self.category_entry = tk.Entry(top_frame, width=30)
        self.category_entry.grid(row=2, column=1, padx=5, pady=(5, 0), sticky="w")

        # Radiobuttons voor persoonsnaam (EN/OF)
        tk.Label(top_frame, text="Persoonsfilter:").grid(row=3, column=0, sticky="w", pady=(5, 0))
        self.person_mode = tk.StringVar(value="AND")
        tk.Radiobutton(top_frame, text="EN", variable=self.person_mode, value="AND").grid(row=3, column=1, sticky="w")
        tk.Radiobutton(top_frame, text="OF", variable=self.person_mode, value="OR").grid(row=3, column=1, padx=60, sticky="w")

        # Radiobuttons voor categorie (EN/OF)
        tk.Label(top_frame, text="Categoriefilter:").grid(row=4, column=0, sticky="w", pady=(5, 0))
        self.category_mode = tk.StringVar(value="AND")
        tk.Radiobutton(top_frame, text="EN", variable=self.category_mode, value="AND").grid(row=4, column=1, sticky="w")
        tk.Radiobutton(top_frame, text="OF", variable=self.category_mode, value="OR").grid(row=4, column=1, padx=60, sticky="w")

        tk.Button(top_frame, text="Zoeken", command=self.search_photos).grid(row=5, column=2, padx=5, pady=(5, 0))

        # Scrollbaar gebied voor foto's
        self.canvas = Canvas(root)
        self.scrollbar = Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = Frame(self.canvas)

        # Alleen scrollen als de muis boven de canvas hangt
        self.canvas.bind("<Enter>", lambda e: self.canvas.focus_set())
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)   # Windows / macOS
        self.canvas.bind("<Button-4>", self._on_mousewheel)     # Linux omhoog
        self.canvas.bind("<Button-5>", self._on_mousewheel)     # Linux omlaag



        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # Referenties naar PhotoImage objecten bewaren (anders worden ze weggegooid)
        self.photo_refs = []

    def match_filter(self, text, filters, mode):
        if not filters:
            return True
        if mode == "AND":
            return all(f in text for f in filters)
        else:  # OR
            return any(f in text for f in filters)

    def _on_mousewheel(self, event):
        # Windows / macOS
        if event.delta != 0:
            self.canvas.yview_scroll(int(-event.delta / 120), "units")
        # Linux
        elif event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")


    def show_full_image(self, img_path):
        top = tk.Toplevel(self.root)
        top.title(os.path.basename(img_path))
        top.state("zoomed")   # Windows: maximaliseer venster

        # Canvas + scrollbar voor grote foto's
        canvas = tk.Canvas(top, bg="black")
        canvas.pack(fill="both", expand=True)

        # Laad de afbeelding
        img = Image.open(img_path)
        photo = ImageTk.PhotoImage(img)

        # Bewaar referentie
        canvas.image = photo

        # Plaats afbeelding in canvas
        canvas.create_image(0, 0, anchor="nw", image=photo)

        # Pas canvasgrootte aan
        canvas.config(scrollregion=canvas.bbox("all"))



        # Scrollbars toevoegen
        vbar = tk.Scrollbar(top, orient="vertical", command=canvas.yview)
        hbar = tk.Scrollbar(top, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)

        vbar.pack(side="right", fill="y")
        hbar.pack(side="bottom", fill="x")

        # Scrollwiel in popup
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta/120), "units"))



    def browse_excel(self):
        path = filedialog.askopenfilename(
            title="Kies Excel-bestand",
            filetypes=[("Excel bestanden", "*.xlsx *.xlsm *.xltx *.xltm"), ("Alle bestanden", "*.*")]
        )
        if path:
            self.excel_entry.delete(0, tk.END)
            self.excel_entry.insert(0, path)

    def search_photos(self):
        excel_path = self.excel_entry.get().strip()
        person_input = self.person_entry.get().strip()
        category_input = self.category_entry.get().strip()

        if not excel_path or not os.path.isfile(excel_path):
            messagebox.showerror("Fout", "Kies een geldig Excel-bestand.")
            return

        # Meerdere waarden verwerken
        def split_values(text):
            text = text.replace(";", ",").replace(" ", ",")
            return [t.strip().lower() for t in text.split(",") if t.strip()]

        person_filters = split_values(person_input)
        category_filters = split_values(category_input)

        if not person_filters and not category_filters:
            messagebox.showerror("Fout", "Vul een persoonsnaam en/of categorie in.")
            return

        # Clear vorige resultaten
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.photo_refs.clear()

        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Fout", f"Kon Excel niet openen:\n{e}")
            return

        base_dir = os.path.dirname(excel_path)
        matches = []

        for row in ws.iter_rows(min_row=2):
            file_cell = row[COLUMN_FILE - 1]
            person_cell = row[COLUMN_PERSON - 1]
            category_cell = row[4]  # Kolom E

            file_value = file_cell.value
            if not file_value:
                continue

            person_text = str(person_cell.value).lower() if person_cell.value else ""
            category_text = str(category_cell.value).lower() if category_cell.value else ""

            # --- FILTERS ---
            match_person = self.match_filter(person_text, person_filters, self.person_mode.get())
            match_category = self.match_filter(category_text, category_filters, self.category_mode.get())

            if not (match_person and match_category):
                continue
            # ----------------

            # Hyperlink of gewone tekst?
            if file_cell.hyperlink and file_cell.hyperlink.target:
                img_path = file_cell.hyperlink.target
            else:
                img_path = str(file_value)

            if not os.path.isabs(img_path):
                img_path = os.path.join(base_dir, img_path)

            matches.append(img_path)

        if not matches:
            messagebox.showinfo("Geen resultaten", "Geen foto's gevonden die voldoen aan de filters.")
            return

        # Foto's tonen
        row_idx = 0
        for img_path in matches:
            if not os.path.isfile(img_path):
                lbl = tk.Label(self.scrollable_frame, text=f"Bestand niet gevonden: {img_path}", fg="red")
                lbl.grid(row=row_idx, column=0, sticky="w", padx=10, pady=5)
                row_idx += 1
                continue

            try:
                img = Image.open(img_path)
                img.thumbnail((400, 400))
                photo = ImageTk.PhotoImage(img)
                self.photo_refs.append(photo)

                frame = tk.Frame(self.scrollable_frame, borderwidth=1, relief="solid")
                frame.grid(row=row_idx, column=0, padx=10, pady=10, sticky="w")

                tk.Label(frame, text=os.path.basename(img_path)).pack(anchor="w", padx=5, pady=2)
                img_label = tk.Label(frame, image=photo, cursor="hand2")
                img_label.pack(padx=5, pady=5)

                # Klik-event koppelen
                img_label.bind("<Button-1>", lambda e, p=img_path: self.show_full_image(p))


                row_idx += 1
            except Exception as e:
                lbl = tk.Label(self.scrollable_frame, text=f"Fout bij laden van {img_path}: {e}", fg="red")
                lbl.grid(row=row_idx, column=0, sticky="w", padx=10, pady=5)
                row_idx += 1




if __name__ == "__main__":
    root = tk.Tk()
    app = PhotoSearchApp(root)
    root.mainloop()
