import importlib
import os
import re
import sys
import pickle
import threading
import tkinter as tk
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
from tkinter import messagebox

try:
    keras_app = importlib.import_module("tensorflow.keras.applications.resnet50")
    keras_image = importlib.import_module("tensorflow.keras.preprocessing.image")
except ModuleNotFoundError:
    try:
        keras_app = importlib.import_module("keras.applications.resnet50")
        keras_image = importlib.import_module("keras.preprocessing.image")
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Keras/TensorFlow is niet geïnstalleerd. Installeer met: pip install tensorflow keras"
        ) from exc

ResNet50 = keras_app.ResNet50
preprocess_input = keras_app.preprocess_input
decode_predictions = keras_app.decode_predictions
image = keras_image

try:
    import cv2
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError("OpenCV is niet geïnstalleerd. Installeer met: pip install opencv-python") from exc

try:
    import numpy as np
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError("NumPy is niet geïnstalleerd. Installeer met: pip install numpy") from exc

try:
    from deepface import DeepFace
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError("DeepFace is niet geïnstalleerd. Installeer met: pip install deepface") from exc
except ValueError as exc:
    raise ValueError(
        "DeepFace heeft aanvullende TensorFlow-compatibiliteit nodig. Installeer tf-keras met: pip install tf-keras"
    ) from exc

try:
    import mediapipe as mp
    from mediapipe.tasks import python
    from mediapipe.tasks.python import vision
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError("MediaPipe is niet geïnstalleerd. Installeer met: pip install mediapipe") from exc

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError("openpyxl is niet geïnstalleerd. Installeer met: pip install openpyxl") from exc

os.environ['TCL_LIBRARY'] = r'C:\Users\Steamlink\AppData\Local\Programs\Python\Python311\tcl\tcl8.6'
os.environ['TK_LIBRARY'] = r'C:\Users\Steamlink\AppData\Local\Programs\Python\Python311\tcl\tk8.6'
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# ==============================
# CONFIG
# ==============================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# OpenCV data pad instellen voor gecompileerde exe
if getattr(sys, 'frozen', False):
    os.environ['OPENCV_HAARCASCADES'] = os.path.join(sys._MEIPASS, 'cv2', 'data')
INPUT_FOLDER = os.path.join(BASE_DIR, "photos")
TRAIN_FOLDER = os.path.join(BASE_DIR, "faces_train")
UNKNOWN_FOLDER = os.path.join(BASE_DIR, "unknown_faces")
DEBUG_FOLDER = os.path.join(BASE_DIR, "debug_faces")

DB_PATH = os.path.join(BASE_DIR, "face_db.pkl")
CSV_PATH = os.path.join(BASE_DIR, "foto_overzicht.csv")
if getattr(sys, 'frozen', False):
    MODEL_PATH = os.path.join(sys._MEIPASS, "blaze_face_full_range.tflite")
else:
    MODEL_PATH = os.path.join(BASE_DIR, "blaze_face_full_range.tflite")

os.makedirs(UNKNOWN_FOLDER, exist_ok=True)
os.makedirs(DEBUG_FOLDER, exist_ok=True)
model = ResNet50(weights="imagenet")

SELECTED_PERSONS = None

# ==============================
# DATABASE
# ==============================
def load_db():
    if os.path.exists(DB_PATH):
        with open(DB_PATH, "rb") as f:
            db = pickle.load(f)

        # fix oude db zonder mean
        for p in db:
            if "mean" not in db[p]:
                db[p]["mean"] = np.mean(db[p]["embeddings"], axis=0)

        print(f"📂 Database geladen ({len(db)} personen)")
        return db
    else:
        print("📂 Nieuwe database")
        return {}

def save_db(db):
    with open(DB_PATH, "wb") as f:
        pickle.dump(db, f)

face_db = load_db()

# ==============================
# DETECTOR
# ==============================
def create_detector(threshold):
    options =vision.FaceDetectorOptions(
        base_options=python.BaseOptions(model_asset_path=MODEL_PATH),
        min_detection_confidence=threshold
    )
    return vision.FaceDetector.create_from_options(options)

detector = None

# ==============================
# HELPERS
# ==============================
def get_color(name):
    np.random.seed(abs(hash(name)) % 10000)
    return tuple(int(x) for x in np.random.randint(0,255,3))

def find_match(embedding, threshold):
    best_match = None
    best_score = 0

    for person, data in face_db.items():
        if SELECTED_PERSONS and person not in SELECTED_PERSONS:
            continue

        mean = data["mean"]

        sim = np.dot(embedding, mean) / (
            np.linalg.norm(embedding) * np.linalg.norm(mean)
        )
        if sim > best_score:
            best_score = sim
            best_match = person

    if best_score > threshold:
        return best_match, best_score

    return None, best_score

# ==============================
# TRAINING
# ==============================
def train_faces(log=print):
    log("🧠 Training starten...")

    if not os.path.exists(TRAIN_FOLDER):
        log("⚠️ Geen faces_train map")
        return

    face_db.clear()

    for person in os.listdir(TRAIN_FOLDER):
        path = os.path.join(TRAIN_FOLDER, person)
        if not os.path.isdir(path):
            continue

        embeddings = []

        for file in os.listdir(path):
            if not file.lower().endswith((".jpg",".png",".jpeg")):
                continue

            img = cv2.imread(os.path.join(path,file))

            try:
                emb = DeepFace.represent(
                    img,
                    model_name="Facenet512",
                    enforce_detection=False
                )[0]["embedding"]

                embeddings.append(emb)

            except Exception as e:
                log(f"❌ {file}: {e}")

        if embeddings:
            face_db[person] = {
                "embeddings": embeddings,
                "mean": np.mean(embeddings, axis=0)
            }

            log(f"✅ {person} ({len(embeddings)} foto's)")

    save_db(face_db)
    log("✅ Training klaar")

# ==============================
# SCAN
# ==============================
stop_flag = False

def classify_image(img_path):
    img = image.load_img(img_path, target_size=(224,224))
    x = image.img_to_array(img)
    x = np.expand_dims(x, axis=0)
    x = preprocess_input(x)

    preds = model.predict(x, verbose=0)
    labels = decode_predictions(preds, top=3)[0]

    return [(l[1], float(l[2])) for l in labels]

def map_to_categories(label_tuples):
    return [label for label, _ in label_tuples[:3]]

def process_image(img_path, detect_th, match_th):
    img = cv2.imread(img_path)
    if img is None:
        return [], []

    debug_img = img.copy()
    persons = []
    confs = []

    mp_image = mp.Image(
        image_format=mp.ImageFormat.SRGB,
        data=cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    )

    result = detector.detect(mp_image)

    if result.detections:
        for i, det in enumerate(result.detections):
            if stop_flag:
                return [], []

            score = det.categories[0].score
            if score < detect_th:
                continue

            bbox = det.bounding_box
            x,y,w,h = int(bbox.origin_x), int(bbox.origin_y), int(bbox.width), int(bbox.height)

            if w < 40 or h < 40:
                continue

            face = img[y:y+h, x:x+w]

            try:
                emb = DeepFace.represent(face, model_name="Facenet512", enforce_detection=False)[0]["embedding"]

                person, sim = find_match(emb, match_th)

                if person:
                    persons.append(person)
                    confs.append(round(sim,2))
                    color = get_color(person)
                    label = f"{person} ({sim:.2f})"
                else:
                    name = f"unknown_{os.path.basename(img_path)}_{i}.jpg"
                    cv2.imwrite(os.path.join(UNKNOWN_FOLDER,name), face)
                    color = (0,0,255)
                    label = "UNKNOWN"

                cv2.rectangle(debug_img,(x,y),(x+w,y+h),color,2)
                cv2.rectangle(debug_img,(x,y-25),(x+200,y),(0,0,0),-1)
                cv2.putText(debug_img,label,(x,y-5),
                            cv2.FONT_HERSHEY_SIMPLEX,0.6,(255,255,255),2)

            except Exception as e:
                print("DeepFace fout:",e)

        cv2.imwrite(os.path.join(DEBUG_FOLDER, os.path.basename(img_path)), debug_img)

    return persons, confs

def save_workbook(wb, excel_path, log):
    while True:
        try:
            wb.save(excel_path)
            return True
        except PermissionError:
            answer = tk.messagebox.askretrycancel(
                "Bestand geblokkeerd",
                f"Het bestand '{os.path.basename(excel_path)}' is mogelijk nog geopend.\n\n"
                "Sluit het bestand en klik op Opnieuw, of klik Annuleren om te stoppen."
            )
            if not answer:
                log("❌ Opslaan geannuleerd.")
                return False
        except Exception as e:
            answer = tk.messagebox.askokcancel(
                "Fout bij opslaan",
                f"Onverwachte fout:\n{e}\n\nWil je het toch proberen te overschrijven?"
            )
            if not answer:
                log(f"❌ Opslaan mislukt: {e}")
                return False

def run_scan(log, detect_th, match_th):
    global stop_flag
    stop_flag = False

    data = []
    i = 1

    for file in os.listdir(INPUT_FOLDER):
        if stop_flag:
            log("⛔ Scan gestopt")
            return

        if file.lower().endswith((".jpg",".png",".jpeg")):
            path = os.path.join(INPUT_FOLDER, file)

            log(f"📸 {file}")

            persons, confs = process_image(path, detect_th, match_th)

            label_data = classify_image(path)
            categories = map_to_categories(label_data)

            if persons:
                categories.append("gezichten")

            categories = list(set(categories))

            data.append([
                i,
                file,
                ", ".join(persons),
                ", ".join(map(str, confs)),
                ", ".join(categories)
            ])

            i += 1

    # Excel schrijven
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["nr", "bestand", "personen", "confidence", "categorieen"]
    ws.append(headers)

    # Pak de header-rij (meestal rij 1)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)

    for row in data:
        nr, bestand, personen, confidence, categorieen = row
        ws.append([nr, bestand, personen, confidence, categorieen])
        cell = ws.cell(row=ws.max_row, column=2)
        file_path = os.path.join(INPUT_FOLDER, bestand)
        cell.hyperlink = file_path
        cell.style = "Hyperlink"

    for col_num, column_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

    folder_name = os.path.basename(INPUT_FOLDER)
    excel_path = os.path.normpath(os.path.join(INPUT_FOLDER, f"foto_overzicht_{folder_name}.xlsx"))

    if save_workbook(wb, excel_path, log):
        log(f"✅ Scan klaar, resultaten opgeslagen in {excel_path}")

# Controle op jaartal aan het eind van foldernaam
def is_year_folder(folder_name):
    match = re.search(r'(\d{4})$', folder_name)
    if match:
        year = int(match.group(1))
        return 1900 <= year <= 2050
    return False

def run_scan_multiple_folders(log, detect_th, match_th, base_folder):
    global stop_flag
    stop_flag = False

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # verwijder standaard leeg blad

    if not is_year_folder(os.path.basename(base_folder)):
        log(f"De map '{base_folder}' eindigt niet op een jaar tussen 1900 en 2050. Scan stopt.")
        return

    subfolders = [f.path for f in os.scandir(base_folder) if f.is_dir()]

    if not subfolders:
        log("Geen submappen gevonden in de gekozen map.")
        return

    for subfolder in subfolders:
        if stop_flag:
            log("⛔ Scan gestopt")
            break

        log(f"📁 Scannen van submap: {os.path.basename(subfolder)}")

        data = []
        i = 1
        for file in os.listdir(subfolder):
            if stop_flag:
                break

            if file.lower().endswith((".jpg", ".png", ".jpeg")):
                path = os.path.join(subfolder, file)

                log(f"📸 {file}")

                persons, confs = process_image(path, detect_th, match_th)

                label_data = classify_image(path)
                categories = map_to_categories(label_data)

                if persons:
                    categories.append("gezichten")

                categories = list(set(categories))

                data.append([
                    i,
                    file,
                    ", ".join(persons),
                    ", ".join(map(str, confs)),
                    ", ".join(categories)
                ])

                i += 1

        ws = wb.create_sheet(title=os.path.basename(subfolder)[:31])

        headers = ["nr", "bestand", "personen", "confidence", "categorieen"]
        ws.append(headers)

        for row in data:
            nr, bestand, personen, confidence, categorieen = row
            ws.append([nr, bestand, personen, confidence, categorieen])
            cell = ws.cell(row=ws.max_row, column=2)
            file_path = os.path.join(subfolder, bestand)
            cell.hyperlink = file_path
            cell.style = "Hyperlink"

        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20

        ws.column_dimensions["A"].width = 7  # ongeveer 50 pixels
        ws.column_dimensions["B"].width = 57  # ongeveer 400 pixels

    folder_name = os.path.basename(base_folder)
    excel_path = os.path.normpath(os.path.join(base_folder, f"foto_overzicht_{folder_name}.xlsx"))
    if save_workbook(wb, excel_path, log):
        log(f"✅ Scan klaar, resultaten opgeslagen in {excel_path}")


# ==============================
# GUI
# ==============================
def start_gui():

    def log(msg):
        log_box.insert(tk.END, msg + "\n")
        log_box.see(tk.END)
        root.update()

    def select_folder():
        folder = filedialog.askdirectory()
        if folder:
            input_var.set(folder)


    def load_persons():
        person_listbox.delete(0, tk.END)
        all_persons = list(face_db.keys())
        upper_names = [p for p in all_persons if p.isupper()]
        other_names = [p for p in all_persons if not p.isupper()]
        # Sorteer beide lijsten alfabetisch
        upper_names.sort()
        other_names.sort()
        sorted_persons = upper_names + other_names
        for p in sorted_persons:
            person_listbox.insert(tk.END, p)


    def start():
        global INPUT_FOLDER, detector, SELECTED_PERSONS

        INPUT_FOLDER = input_var.get()
        detect_th = float(detect_var.get())
        match_th = float(match_var.get())

        detector = create_detector(detect_th)

        selected = [person_listbox.get(i) for i in person_listbox.curselection()]

        if selected:
            SELECTED_PERSONS = set(selected)
            log(f"👤 Filter: {selected}")
        else:
            SELECTED_PERSONS = None
            log("👤 Alle personen")

        if train_var.get():
            train_faces(log)
            load_persons()

        if is_year_folder(os.path.basename(INPUT_FOLDER)):
            threading.Thread(target=run_scan_multiple_folders, args=(log, detect_th, match_th, INPUT_FOLDER), daemon=True).start()
        else:
            threading.Thread(target=run_scan, args=(log, detect_th, match_th), daemon=True).start()

    def stop():
        global stop_flag
        stop_flag = True

    root = tk.Tk()
    root.title("AI Foto Scanner")

    input_var = tk.StringVar(value=INPUT_FOLDER)
    detect_var = tk.StringVar(value="0.5")
    match_var = tk.StringVar(value="0.6")
    train_var = tk.BooleanVar()

    tk.Label(root,text="Map").grid(row=0,column=0)
    tk.Entry(root,textvariable=input_var,width=50).grid(row=0,column=1)
    tk.Button(root,text="Bladeren", command=select_folder).grid(row=0,column=2)

    tk.Label(root,text="Detect threshold").grid(row=1,column=0)
    tk.Entry(root,textvariable=detect_var).grid(row=1,column=1)

    tk.Label(root,text="Match threshold").grid(row=2,column=0)
    tk.Entry(root,textvariable=match_var).grid(row=2,column=1)

    tk.Button(root,text="🚀 Start",command=start).grid(row=4,column=0)
    tk.Button(root,text="⛔ Stop",command=stop).grid(row=4,column=1)
    tk.Button(root,text="🧠 Train",command=lambda: threading.Thread(target=train_faces, args=(log,), daemon=True).start()).grid(row=4,column=2)

    tk.Label(root, text="Selecteer personen (leeglaten = alle personen)").grid(row=5, column=0, columnspan=3)

    person_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE, height=8, width=40)
    person_listbox.grid(row=6, column=0, columnspan=3)

    load_persons()

    log_box = ScrolledText(root,width=80,height=15)
    log_box.grid(row=7,column=0,columnspan=3)

    root.mainloop()

# ==============================
if __name__ == "__main__":
    start_gui()
